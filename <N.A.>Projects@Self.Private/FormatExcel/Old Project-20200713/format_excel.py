from openpyxl import load_workbook
import re

wb = load_workbook('Data/source.xlsx')
working_sheet = wb["1"]
num_units = working_sheet['A']
old_unit_price = working_sheet['B']
old_quantity = working_sheet['C']
old_amount = working_sheet['D']


def extract_unit(num_unit):
    num_unit = str(num_unit)
    pattern = re.compile("[\u4e00-\u9fa5]")
    return "".join(pattern.findall(num_unit))


def extract_num(num_unit):
    num_unit = str(num_unit)
    pattern = re.compile("[0-9]")
    num = "".join(pattern.findall(num_unit))
    if num:
        return num
    else:
        return 1


for index, item in enumerate(num_units):
    item_unit = extract_unit(item.value)
    item_fix_coefficient = float(extract_num(item.value))
    try:
        working_sheet.cell(row=index+1, column=6).value = item_unit
    except:
        continue
    try:
        working_sheet.cell(row=index+1, column=7).value = old_unit_price[index].value * item_fix_coefficient
    except:
        continue
    try:
        working_sheet.cell(row=index+1, column=8).value = old_quantity[index].value / item_fix_coefficient
    except:
        continue
    try:
        working_sheet.cell(row=index+1, column=9).value = \
            working_sheet.cell(row=index+1, column=7).value * working_sheet.cell(row=index+1, column=8).value
    except:
        continue
    try:
        working_sheet.cell(row=index+1, column=10).value = \
            working_sheet.cell(row=index+1, column=9).value - old_amount[index].value
    except:
        continue

wb.save('output.xlsx')
