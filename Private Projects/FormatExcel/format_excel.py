from openpyxl import load_workbook
num_columns = 33  # --------------------Number of Rows to copy-------------------- #


#%%
# Main Logic 1 - Find the relationship between different ranks of indices.
def get_related_rank_iii_index(_index, rank_iii_index_list):
    for j in range(len(rank_iii_index_list)):
        if j < len(rank_iii_index_list) - 1:
            if (_index > rank_iii_index_list[j]) and (_index < rank_iii_index_list[j+1]):
                return rank_iii_index_list[j]
            else:
                continue
        else:
            return rank_iii_index_list[j]


def get_related_rank_ii_index(_index, rank_ii_index_list):
    for j in range(len(rank_ii_index_list)):
        if j < len(rank_ii_index_list) - 1:
            if (_index > rank_ii_index_list[j]) and (_index < rank_ii_index_list[j+1]):
                return rank_ii_index_list[j]
            else:
                continue
        else:
            return rank_ii_index_list[j]


def get_related_rank_i_index(_index, rank_i_index_list):
    for j in range(len(rank_i_index_list)):
        if j < len(rank_i_index_list) - 1:
            if (_index > rank_i_index_list[j]) and (_index < rank_i_index_list[j+1]):
                return rank_i_index_list[j]
            else:
                continue
        else:
            return rank_i_index_list[j]


def get_all_rank_indices(_index, _rank_i_indices, _rank_ii_indices, _rank_iii_indices):
    _rank_iii_index = get_related_rank_iii_index(_index, _rank_iii_indices)
    _rank_ii_index = get_related_rank_ii_index(_rank_iii_index, _rank_ii_indices)
    _rank_i_index = get_related_rank_i_index(_rank_ii_index, _rank_i_indices)
    return _rank_i_index, _rank_ii_index, _rank_iii_index, _index


def get_final_indices(divided_section_index):
    divided_section_matrix = []
    for _index in divided_section_index:
        _a, _b, _c, _d = get_all_rank_indices(_index, rank_I_index, rank_II_index, rank_III_index)
        divided_section_matrix.append([_a, _b, _c, _d])
    final_list = []
    for k in range(len(divided_section_matrix)):
        list_k = divided_section_matrix[k]
        final_list = list(set(list_k + final_list))
    final_list.sort()
    return final_list


def write_file(_final_indices, destination_sheet_name):
    def copy_rows(divided_row_indices):
        _copied_range = []
        for _row in divided_row_indices:
            _copied_row = []
            for _column in range(1, num_columns + 1):
                _copied_row.append(ws.cell(row=_row, column=_column).value)
            _copied_range.append(_copied_row)
        return _copied_range

    _ws = wb[destination_sheet_name]
    _temp = copy_rows(_final_indices)
    for _row in range(1, len(_temp) + 1):
        for _column in range(1, num_columns + 1):
            old_cell = _temp[_row - 1][_column - 1]
            new_cell = _ws.cell(row=_row, column=_column)
            new_cell.value = old_cell
            #
            # if _ws[f"{_row - 1}{_column - 1}"].has_style:
            #     new_cell.border = copy(old_cell.border)
            #     new_cell.fill = copy(old_cell.fill)
            #     new_cell.number_format = copy(old_cell.number_format)
            #     new_cell.alignment = copy(old_cell.alignment)


#%%
wb = load_workbook("1.xlsx", data_only=True)
ws = wb["分部分项"]
ws_beam_factory_1 = wb.create_sheet("1#梁场")
ws_beam_factory_2 = wb.create_sheet("2#梁场")
ws_steel_factory_1 = wb.create_sheet("1#钢筋加工棚")
ws_steel_factory_2 = wb.create_sheet("2#钢筋加工棚")
ws_concrete_factory_1 = wb.create_sheet("1#拌合站")
ws_concrete_factory_2 = wb.create_sheet("2#拌合站")
ws_working_area_1 = wb.create_sheet("一工区")
ws_working_area_2 = wb.create_sheet("二工区")
ws_working_area_3 = wb.create_sheet("三工区")

rank_I_index = []
rank_II_index = []
rank_III_index = []
main_contents = []
for i in range(1, ws.max_row + 1):
    if ws[f'A{i}'].value == 'I':
        rank_I_index.append(i)
    elif ws[f'A{i}'].value == 'II':
        rank_II_index.append(i)
    elif ws[f'A{i}'].value == 'III':
        rank_III_index.append(i)
    else:
        main_contents.append(i)

beam_factory_1 = []
beam_factory_2 = []
steel_factory_1 = []
steel_factory_2 = []
concrete_factory_1 = []
concrete_factory_2 = []
working_area_1 = []
working_area_2 = []
working_area_3 = []
for main_content in main_contents:
    if ws[f'D{main_content}'].value == '1#梁场':
        beam_factory_1.append(main_content)
    elif ws[f'D{main_content}'].value == '2#梁场':
        beam_factory_2.append(main_content)
    elif ws[f'D{main_content}'].value == '1#钢筋加工棚':
        steel_factory_1.append(main_content)
    elif ws[f'D{main_content}'].value == '2#钢筋加工棚':
        steel_factory_2.append(main_content)
    elif ws[f'D{main_content}'].value == '一工区':
        working_area_1.append(main_content)
    elif ws[f'D{main_content}'].value == '二工区':
        working_area_2.append(main_content)
    elif ws[f'D{main_content}'].value == '三工区':
        working_area_3.append(main_content)
for main_content in main_contents:
    if ws[f'C{main_content}'].value == '1#拌合站':
        concrete_factory_1.append(main_content)
    elif ws[f'C{main_content}'].value == '2#拌合站':
        concrete_factory_2.append(main_content)

new_sheets = [beam_factory_1, beam_factory_2, steel_factory_1, steel_factory_2, concrete_factory_1, concrete_factory_2,
              working_area_1, working_area_2, working_area_3]
for index, new_sheet in enumerate(new_sheets):
    new_sheets[index] = get_final_indices(new_sheet)

new_sheets_name = ["1#梁场", "2#梁场", "1#钢筋加工棚", "2#钢筋加工棚", "1#拌合站", "2#拌合站", "一工区", "二工区", "三工区"]

for sheet, sheet_name in zip(new_sheets, new_sheets_name):
    write_file(sheet, sheet_name)

wb.save("new.xlsx")
